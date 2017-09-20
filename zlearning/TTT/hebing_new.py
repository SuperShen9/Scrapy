# -*- coding: utf-8 -*-
import os,openpyxl,xlrd
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font
os.chdir('D:\zlianxi\hebing_excell')
fath=unicode('D:\zlianxi\hebing_excell','utf-8')
for x,y,z, in os.walk(fath):
    for y1 in y:
        path1='D:\zlianxi\hebing_excell\\'+str(y1)
        print path1
        os.chdir(path1)
        file = 'baocun.xlsx'
        if os.path.exists(file):
            os.remove(file)
        spam={'ID':2,'name':3,'age':4}
        k1=0
        fath=unicode(path1,'utf-8')
        for foldername,subfolder,excels in os.walk(fath):
            baocun = openpyxl.Workbook()
            sheet = baocun.create_sheet(index=0, title='data')
            for excel in excels:
                wb=xlrd.open_workbook(excel)
                sheet1 = wb.sheets()[0]
                hang = sheet1.nrows
                lie = sheet1.ncols
                for k in range(lie):
                    if sheet1.cell(0,k).value in spam.keys():
                        kk = get_column_letter(spam[sheet1.cell(0,k).value])
                        sheet[kk + '1']=sheet1.cell(0,k).value
                        sheet['A1'] = '来源'
                        fontobj = Font(name='Arial', size=12, bold=True)
                        sheet['A1'].font = fontobj
                        sheet[kk + '1'].font = fontobj
                        j = 2
                        for i in range(1,hang):
                            sheet['A'+str(j+k1)] = excel
                            sheet[kk+str(j+k1)] =sheet1.cell(i,k).value
                            j+=1
                k1+=hang-1
        sheet.freeze_panes='A2'
        baocun.remove_sheet(baocun.get_sheet_by_name('Sheet'))
        baocun.save('baocun.xlsx')