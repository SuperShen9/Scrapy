# -*- coding: utf-8 -*-
# author:Super
import os,openpyxl,pprint
# from TTT import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
os.chdir('D:\zlianxi\New_templete_clean')
wbf = openpyxl.load_workbook('Templete.xlsx')
sheet_tem = wbf.get_sheet_by_name('Sheet1')
hang_tem = sheet_tem.max_row + 1
segment={}
AM={}
job={}
dept={}
indu={}
in_pro={}
b_time={}
b_bug={}
for row in range(2,hang_tem):
    if sheet_tem['A' + str(row)].value != None:
        key_o = sheet_tem['A' + str(row)].value
        val_o = sheet_tem['B' + str(row)].value
        segment.setdefault(key_o, val_o)
    if sheet_tem['C' + str(row)].value != None:
        key_o = sheet_tem['C' + str(row)].value
        val_o = sheet_tem['D' + str(row)].value
        AM.setdefault(key_o, val_o)
    if sheet_tem['E' + str(row)].value != None:
        key_o = sheet_tem['E' + str(row)].value
        val_o = sheet_tem['F' + str(row)].value
        job.setdefault(key_o, val_o)
    if sheet_tem['G' + str(row)].value != None:
        key_o = sheet_tem['G' + str(row)].value
        val_o = sheet_tem['H' + str(row)].value
        dept.setdefault(key_o, val_o)
    if sheet_tem['I' + str(row)].value != None:
        key_o = sheet_tem['I' + str(row)].value
        val_o = sheet_tem['J' + str(row)].value
        indu.setdefault(key_o, val_o)
    if sheet_tem['K' + str(row)].value != None:
        key_o = sheet_tem['K' + str(row)].value
        val_o = sheet_tem['L' + str(row)].value
        in_pro.setdefault(key_o, val_o)
    if sheet_tem['N' + str(row)].value != None:
        key_o = sheet_tem['N' + str(row)].value
        val_o = sheet_tem['O' + str(row)].value
        b_time.setdefault(key_o, val_o)
    if sheet_tem['P' + str(row)].value != None:
        key_o = sheet_tem['P' + str(row)].value
        val_o = sheet_tem['Q' + str(row)].value
        b_bug.setdefault(key_o, val_o)


os.chdir('D:\zlianxi\New_templete_clean\clean')
file = 'Clean_data.xlsx'
if os.path.exists(file):
    os.remove(file)
filepath = unicode('D:\zlianxi\New_templete_clean\clean', 'utf-8')
list1=['Segment']
list2=['AM']
list3=[u'職稱']
list4=[u'部門']
list5=[u'產業別']
list6=[u'有興趣投資的IT解決方案?(可複選)']
list7=[u'專案時程']
list8=[u'專案預算(USD)']
list9=[u'姓名']
list10=[u'完整公司名稱']


for foldername,subfolder,excels in os.walk(filepath):
    Clean_data = openpyxl.Workbook()
    sheet = Clean_data.create_sheet(index=0, title='data')
    wb=openpyxl.load_workbook(excels[0])
    sheet1 = wb.active
    hang = sheet1.max_row+1
    lie = sheet1.max_column+1
    print '数据量统计：%s'%(hang-2)
    i=0
    for k in range(1, lie):
        liebiao = get_column_letter(k)
        liebiao1=get_column_letter(k+i)
        i+=1
        for j in range(1, hang):
            sheet[liebiao1 + str(j)] = sheet1[liebiao + str(j)].value

    hang1 = sheet.max_row + 1
    lie1 = sheet.max_column + 1
    for kk in range(1 , lie1):
        lb=get_column_letter(kk)
        lb_1=get_column_letter(kk+1)
        for jj in range(2,hang1):
            if sheet[lb + '1'].value in list1:
                sheet[lb_1 + '1'] = '标准segment'
                sheet[lb_1 + str(jj)] = segment.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list2:
                sheet[lb_1 + '1'] = '检验AM'
                if sheet[lb + str(jj)].value!=None :
                    if sheet[lb + str(jj)].value in AM:
                        sheet[lb_1 + str(jj)] = AM.get(sheet[lb + str(jj)].value)
                    else:
                        sheet[lb_1 + str(jj)] ='XXXX'
            if sheet[lb + '1'].value in list3:
                sheet[lb_1 + '1'] = '标准職稱'
                sheet[lb_1 + str(jj)] = job.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list4:
                sheet[lb_1 + '1'] = '标准部門'
                sheet[lb_1 + str(jj)] = dept.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list5:
                sheet[lb_1 + '1'] = '标准行业'
                sheet[lb_1 + str(jj)] = indu.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list6:
                sheet[lb_1 + '1'] = '标准产品'
                sheet[lb_1 + str(jj)] = in_pro.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list7:
                sheet[lb_1 + '1'] = '标准时间'
                sheet[lb_1 + str(jj)] = b_time.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list8:
                sheet[lb_1 + '1'] = '标准金额'
                sheet[lb_1 + str(jj)] = b_bug.get(sheet[lb + str(jj)].value)
            if sheet[lb + '1'].value in list9:
                sheet[lb_1 + '1'] = '标准姓名'
                sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value.strip()
            if sheet[lb + '1'].value in list10:
                sheet[lb_1 + '1'] = '标准公司名稱'
                sheet[lb_1 + str(jj)] = sheet[lb + str(jj)].value.strip()


        # sheet[liebiao1 + '1'] = sheet1[liebiao + '1'].value
        # for j in range(2, hang):
        #     if sheet1[liebiao + '1'].value=='COMPANY_NAME' or \
        #         sheet1[liebiao + '1'].value == 'LAST_NAME':
        #         sheet[liebiao + str(j)] = sheet1[liebiao + str(j)].value.strip()
        #
        #     else:
        #         sheet[liebiao + str(j)]=sheet1[liebiao + str(j)].value
        # if sheet1[liebiao + '1'].value.lower() in spam.keys():
        #     kk = get_column_letter(spam[sheet1[liebiao + '1'].value.lower()])
        #     sheet[kk + '1'] = sheet1[liebiao + '1'].value
        #     for j in range(1, hang):
        #         sheet['A' + str(j)] = excels[0]
        #         sheet[kk + str(j)] = sheet1[liebiao + str(j)].value
        #         j += 1
Clean_data.save('Clean_data.xlsx')